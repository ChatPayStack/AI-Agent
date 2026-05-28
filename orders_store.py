# orders_store.py
from datetime import datetime, timezone
from typing import Dict, Any
import os

import openpyxl
from openpyxl import load_workbook

from db import orders


EXCEL_PATH = os.path.join(os.path.dirname(__file__), "orders.xlsx")
HEADERS = ["timestamp", "order_id", "thread_id", "customer", "items", "total", "payment_method", "status"]


def _now():
    return datetime.now(timezone.utc)


def _format_items(cart_snapshot: Any) -> str:
    if not cart_snapshot:
        return ""
    items = cart_snapshot.get("items") or []
    parts = []
    for it in items:
        name = it.get("name", "Unknown")
        qty = it.get("qty", 1)
        parts.append(f"{name} x{qty}")
    return ", ".join(parts)


def _ensure_workbook() -> openpyxl.Workbook:
    if os.path.exists(EXCEL_PATH):
        return load_workbook(EXCEL_PATH)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(HEADERS)
    wb.save(EXCEL_PATH)
    return wb


def append_order_to_excel(order_data: Dict[str, Any]):
    try:
        wb = _ensure_workbook()
        ws = wb.active

        cart_snapshot = order_data.get("cart_snapshot") or {}
        amount = order_data.get("amount", "")
        currency = order_data.get("currency", "")
        total = f"{amount} {currency}".strip() if amount else ""

        payment_id = order_data.get("payment_id")
        payment_method = "stripe" if order_data.get("stripe_session_id") else (
            "crypto" if order_data.get("tx_hash") or order_data.get("tx_signature") else "unknown"
        )

        row = [
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            str(order_data.get("_id", "")),
            str(order_data.get("thread_id", "")),
            str(order_data.get("email", "")),
            _format_items(cart_snapshot),
            total,
            payment_method,
            order_data.get("status", "pending"),
        ]

        ws.append(row)
        wb.save(EXCEL_PATH)
        print(f"✅ Order added to orders.xlsx")
    except Exception as e:
        print(f"❌ Failed to write order to Excel: {e}")


# CREATE order
async def create_order(state: Dict[str, Any]) -> Dict[str, Any]:
    state["status"] = "pending"
    state["created_at"] = _now()
    state["updated_at"] = _now()

    res = await orders(state["business_id"]).insert_one(state)
    state["_id"] = res.inserted_id

    append_order_to_excel(state)

    return state


# LOAD order
async def load_order(order_id, business_id: str):
    return await orders(business_id).find_one({"_id": order_id})


# MARK fulfilled
async def mark_order_fulfilled(order_id, business_id: str):
    await orders(business_id).update_one(
        {"_id": order_id},
        {
            "$set": {
                "status": "fulfilled",
                "updated_at": _now(),
            }
        },
    )
